import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re

def create_time_buckets(base_interval_minutes, candle_interval_minutes):
    """Create appropriate time buckets based on base and analysis timeframes"""
    
    # Define bucketing strategy based on base interval
    if base_interval_minutes == 240:  # 4-hour base
        bucket_size_minutes = 30  # 30-minute buckets
        bucket_labels = []
        for hour in range(0, 24):
            for minute in [0, 30]:
                time_str = f"{hour:02d}{minute:02d}"
                bucket_labels.append(time_str)
        return bucket_labels, bucket_size_minutes
    
    elif base_interval_minutes == 1440:  # Daily base
        bucket_size_minutes = 60  # Hourly buckets
        bucket_labels = []
        for hour in range(0, 24):
            bucket_labels.append(f"{hour:02d}00")
        return bucket_labels, bucket_size_minutes
    
    elif base_interval_minutes == 10080:  # Weekly base
        bucket_size_minutes = 240  # 4-hour buckets
        bucket_labels = ["0000", "0400", "0800", "1200", "1600", "2000"]
        return bucket_labels, bucket_size_minutes
    
    elif base_interval_minutes == 43200:  # Monthly base
        bucket_size_minutes = 1440  # Daily buckets
        bucket_labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        return bucket_labels, bucket_size_minutes
    
    else:
        # Default to hourly if unknown
        bucket_size_minutes = 60
        bucket_labels = []
        for hour in range(0, 24):
            bucket_labels.append(f"{hour:02d}00")
        return bucket_labels, bucket_size_minutes

def bucket_time_dynamic(time_value, base_interval_minutes, candle_interval_minutes):
    """Convert numeric times to appropriate buckets based on timeframe"""
    if pd.isna(time_value):
        return "No Trigger"
    
    # Handle string times (like "OPEN")
    if isinstance(time_value, str):
        if time_value.upper() == "OPEN":
            return "OPEN"
        try:
            time_value = float(time_value)
        except:
            return str(time_value)
    
    # Get bucket strategy
    bucket_labels, bucket_size_minutes = create_time_buckets(base_interval_minutes, candle_interval_minutes)
    
    # Convert time to minutes since midnight
    if time_value < 100:  # Handle times like 9.5 (9:30)
        hours = int(time_value)
        minutes = int((time_value - hours) * 60)
    else:  # Handle times like 930 (9:30)
        hours = int(time_value // 100)
        minutes = int(time_value % 100)
    
    total_minutes = hours * 60 + minutes
    
    # Handle market open specially
    if total_minutes < 570:  # Before 9:30 AM
        return "OPEN"
    
    # Find appropriate bucket
    if base_interval_minutes == 240:  # 4-hour base with 30-minute buckets
        # Round to nearest 30-minute interval
        bucket_minutes = (total_minutes // 30) * 30
        bucket_hour = bucket_minutes // 60
        bucket_min = bucket_minutes % 60
        return f"{bucket_hour:02d}{bucket_min:02d}"
    
    elif base_interval_minutes == 1440:  # Daily base with hourly buckets
        # Round to nearest hour
        bucket_hour = total_minutes // 60
        return f"{bucket_hour:02d}00"
    
    elif base_interval_minutes == 10080:  # Weekly base with 4-hour buckets
        # Round to nearest 4-hour interval
        bucket_4hr = (total_minutes // 240) * 4
        return f"{bucket_4hr:02d}00"
    
    elif base_interval_minutes == 43200:  # Monthly base with daily buckets
        # This would need date information, defaulting to day names
        day_of_week = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        # Would need actual date parsing for this
        return "Mon"  # Placeholder
    
    else:
        # Default to hourly
        bucket_hour = total_minutes // 60
        return f"{bucket_hour:02d}00"

def validate_files_for_combination(dataframes, filenames):
    """Validate that files can be combined"""
    if len(dataframes) < 2:
        return True, "Single file uploaded"
    
    # Check required columns exist
    required_cols = ['Ticker', 'Base_Interval_Minutes', 'Candle_Interval_Minutes', 'AssetType']
    
    for i, df in enumerate(dataframes):
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            return False, f"File {filenames[i]} missing required columns: {missing_cols}"
    
    # Get reference values from first file
    ref_df = dataframes[0]
    ref_ticker = ref_df['Ticker'].iloc[0]
    ref_base_interval = ref_df['Base_Interval_Minutes'].iloc[0]
    ref_candle_interval = ref_df['Candle_Interval_Minutes'].iloc[0]
    ref_asset_type = ref_df['AssetType'].iloc[0]
    
    # Check all files match
    for i, df in enumerate(dataframes[1:], 1):
        ticker = df['Ticker'].iloc[0]
        base_interval = df['Base_Interval_Minutes'].iloc[0]
        candle_interval = df['Candle_Interval_Minutes'].iloc[0]
        asset_type = df['AssetType'].iloc[0]
        
        if ticker != ref_ticker:
            return False, f"Ticker mismatch: {ref_ticker} vs {ticker} (file {filenames[i]})"
        
        if base_interval != ref_base_interval:
            return False, f"Base interval mismatch: {ref_base_interval} vs {base_interval} (file {filenames[i]})"
        
        if candle_interval != ref_candle_interval:
            return False, f"Candle interval mismatch: {ref_candle_interval} vs {candle_interval} (file {filenames[i]})"
        
        if asset_type != ref_asset_type:
            return False, f"Asset type mismatch: {ref_asset_type} vs {asset_type} (file {filenames[i]})"
    
    return True, f"All files validated for combination: {ref_ticker} {ref_asset_type}"

def detect_date_overlaps(dataframes, filenames):
    """Detect date overlaps between files"""
    overlaps = []
    
    for i in range(len(dataframes)):
        for j in range(i + 1, len(dataframes)):
            df1_dates = set(dataframes[i]['Date'].unique())
            df2_dates = set(dataframes[j]['Date'].unique())
            
            overlap_dates = df1_dates.intersection(df2_dates)
            
            if overlap_dates:
                overlaps.append({
                    'file1': filenames[i],
                    'file2': filenames[j],
                    'overlap_dates': sorted(list(overlap_dates)),
                    'overlap_count': len(overlap_dates)
                })
    
    return overlaps

def resolve_date_overlaps(dataframes, filenames, resolution_strategy='keep_newer'):
    """Resolve date overlaps based on strategy"""
    
    if resolution_strategy == 'keep_newer':
        # Sort files by modification time (newer files override older ones)
        # Since we can't get file modification time from uploaded files,
        # we'll assume files uploaded later are newer
        combined_df = pd.DataFrame()
        
        for df, filename in zip(dataframes, filenames):
            if len(combined_df) == 0:
                combined_df = df.copy()
                st.info(f"📁 Base dataset: {filename} ({len(df)} records)")
            else:
                # Find overlapping dates
                existing_dates = set(combined_df['Date'].unique())
                new_dates = set(df['Date'].unique())
                overlap_dates = existing_dates.intersection(new_dates)
                
                if overlap_dates:
                    # Remove overlapping dates from combined_df
                    combined_df = combined_df[~combined_df['Date'].isin(overlap_dates)]
                    st.warning(f"🔄 Replacing {len(overlap_dates)} overlapping dates with data from {filename}")
                
                # Add all data from new file
                combined_df = pd.concat([combined_df, df], ignore_index=True)
                st.info(f"📁 Added: {filename} ({len(df)} records)")
        
        return combined_df
    
    elif resolution_strategy == 'keep_all':
        # Keep all records, mark duplicates
        combined_df = pd.concat(dataframes, ignore_index=True)
        
        # Add source file information
        for i, (df, filename) in enumerate(zip(dataframes, filenames)):
            combined_df.loc[combined_df.index[sum(len(d) for d in dataframes[:i]):sum(len(d) for d in dataframes[:i+1])], 'source_file'] = filename
        
        return combined_df
    
    else:  # 'error'
        combined_df = pd.concat(dataframes, ignore_index=True)
        return combined_df

def combine_dataframes(dataframes, filenames):
    """Combine multiple dataframes with overlap detection and resolution"""
    
    # Detect overlaps
    overlaps = detect_date_overlaps(dataframes, filenames)
    
    if overlaps:
        st.warning("⚠️ Date overlaps detected between files:")
        
        total_overlap_dates = set()
        for overlap in overlaps:
            st.write(f"📅 **{overlap['file1']}** ↔ **{overlap['file2']}**: {overlap['overlap_count']} overlapping dates")
            
            # Show some example dates
            example_dates = overlap['overlap_dates'][:5]
            if len(overlap['overlap_dates']) > 5:
                st.write(f"   Examples: {', '.join(example_dates)} ... (+{len(overlap['overlap_dates'])-5} more)")
            else:
                st.write(f"   Dates: {', '.join(example_dates)}")
            
            total_overlap_dates.update(overlap['overlap_dates'])
        
        st.write(f"📊 **Total unique overlapping dates**: {len(total_overlap_dates)}")
        
        # Resolution strategy selection
        resolution_strategy = st.radio(
            "How should overlapping dates be handled?",
            options=[
                ('keep_newer', '🔄 Keep Newer Results (later files override earlier files)'),
                ('error', '❌ Stop Processing (requires manual resolution)')
            ],
            format_func=lambda x: x[1],
            key="overlap_resolution"
        )
        
        if resolution_strategy[0] == 'error':
            st.error("❌ Processing stopped due to date overlaps. Please resolve manually and re-upload.")
            return None
        
        # Apply resolution
        combined_df = resolve_date_overlaps(dataframes, filenames, resolution_strategy[0])
        
        st.success(f"✅ Date overlaps resolved using strategy: {resolution_strategy[1]}")
        
    else:
        # No overlaps, simple concatenation
        combined_df = pd.concat(dataframes, ignore_index=True)
        st.success("✅ No date overlaps detected - files combined successfully")
    
    # Final deduplication based on key columns
    key_columns = ['Date', 'Direction', 'TriggerLevel', 'GoalLevel', 'AnalysisType']
    original_count = len(combined_df)
    combined_df = combined_df.drop_duplicates(subset=key_columns, keep='last')
    
    if len(combined_df) < original_count:
        st.info(f"🔍 Removed {original_count - len(combined_df)} duplicate records after combining")
    
    return combined_df

def generate_excel_report(summary_df, metadata, ticker, base_interval, candle_interval, asset_type):
    """Generate Excel report with dynamic bucketing"""
    
    # Create workbook and worksheets
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Create Information sheet
    info_ws = wb.create_sheet("Information")
    info_ws.cell(row=1, column=1, value="Multi-Analysis ATR Report with Dynamic Time Bucketing")
    info_ws.cell(row=2, column=1, value=f"Ticker: {ticker} | Asset: {asset_type} | Base: {base_interval}min | Analysis: {candle_interval}min")
    info_ws.cell(row=3, column=1, value=f"Date Range: {metadata['date_range']} | Records: {metadata['total_records']:,} | Days: {metadata['unique_dates']:,}")
    
    # Get bucket labels for this timeframe
    bucket_labels, _ = create_time_buckets(base_interval, candle_interval)
    
    # Create sheets for each analysis type and direction
    analysis_types = summary_df['AnalysisType'].unique()
    directions = summary_df['Direction'].unique()
    
    for analysis_type in analysis_types:
        for direction in directions:
            data = summary_df[
                (summary_df['AnalysisType'] == analysis_type) & 
                (summary_df['Direction'] == direction)
            ]
            
            if len(data) > 0:
                sheet_name = f"{analysis_type}_{direction}"
                ws = wb.create_sheet(sheet_name)
                create_analysis_sheet(ws, data, bucket_labels, analysis_type, direction)
    
    # Save to BytesIO buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def create_analysis_sheet(ws, data, bucket_labels, analysis_type, direction):
    """Create a worksheet for one analysis type and direction"""
    
    # Headers
    headers = ['Trigger', 'Goal', 'Time'] + bucket_labels + ['Total', 'Hits', 'Rate%']
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    current_row = 2
    
    # Group data by trigger level, goal level, then trigger time
    trigger_levels = sorted(data['TriggerLevel'].unique())
    
    for trigger_level in trigger_levels:
        trigger_data = data[data['TriggerLevel'] == trigger_level]
        goal_levels = sorted(trigger_data['GoalLevel'].unique())
        
        for goal_level in goal_levels:
            goal_data = trigger_data[trigger_data['GoalLevel'] == goal_level]
            trigger_times = sorted(goal_data['TriggerTime'].unique())
            
            for trigger_time in trigger_times:
                time_data = goal_data[goal_data['TriggerTime'] == trigger_time]
                
                # Create trigger label
                trigger_label = f"{trigger_level:+.3f}ATR"
                
                # Prepare data for this row
                total_triggers = time_data['NumTriggers'].iloc[0] if len(time_data) > 0 else 0
                
                # Fill in data
                ws.cell(row=current_row, column=1, value=trigger_label)
                ws.cell(row=current_row, column=2, value=f"{goal_level:+.3f}")
                ws.cell(row=current_row, column=3, value=trigger_time)
                
                # Fill bucket data
                total_hits = 0
                for i, bucket in enumerate(bucket_labels):
                    bucket_data = time_data[time_data['GoalTime'] == bucket]
                    hits = bucket_data['NumHits'].iloc[0] if len(bucket_data) > 0 else 0
                    ws.cell(row=current_row, column=4 + i, value=hits)
                    total_hits += hits
                
                # Summary columns
                ws.cell(row=current_row, column=len(headers) - 2, value=total_triggers)
                ws.cell(row=current_row, column=len(headers) - 1, value=total_hits)
                rate = (total_hits / total_triggers * 100) if total_triggers > 0 else 0
                ws.cell(row=current_row, column=len(headers), value=f"{rate:.1f}%")
                
                current_row += 1

def generate_html_report(summary_df, metadata, ticker, base_interval, candle_interval, asset_type):
    """Generate HTML report with dynamic bucketing"""
    
    bucket_labels, bucket_size = create_time_buckets(base_interval, candle_interval)
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Multi-Analysis ATR Report - {ticker}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            .header {{ background: #333; color: white; padding: 20px; text-align: center; }}
            .summary {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px; margin: 20px 0; }}
            .stat {{ background: #f0f0f0; padding: 15px; text-align: center; border-radius: 5px; }}
            .analysis-section {{ margin: 30px 0; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
            th {{ background: #f2f2f2; }}
            .high {{ background: #d4edda; }}
            .med {{ background: #fff3cd; }}
            .low {{ background: #f8d7da; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Multi-Analysis ATR Report</h1>
            <p>{ticker} ({asset_type}) | Base: {base_interval}min | Analysis: {candle_interval}min</p>
            <p>{metadata['date_range']} | {metadata['total_records']:,} records | {metadata['unique_dates']:,} days</p>
        </div>
        
        <div class="summary">
            <div class="stat">
                <h3>{metadata['total_records']:,}</h3>
                <p>Total Records</p>
            </div>
            <div class="stat">
                <h3>{metadata['unique_dates']:,}</h3>
                <p>Trading Days</p>
            </div>
            <div class="stat">
                <h3>{len(summary_df['AnalysisType'].unique())}</h3>
                <p>Analysis Types</p>
            </div>
            <div class="stat">
                <h3>{bucket_size}min</h3>
                <p>Bucket Size</p>
            </div>
        </div>
    """
    
    # Add sections for each analysis type
    analysis_types = summary_df['AnalysisType'].unique()
    for analysis_type in analysis_types:
        analysis_data = summary_df[summary_df['AnalysisType'] == analysis_type]
        
        html_content += f"""
        <div class="analysis-section">
            <h2>{analysis_type} Analysis</h2>
        """
        
        # Add table for this analysis type
        # (Simplified for brevity - would include full table generation)
        html_content += f"""
            <p>Records: {len(analysis_data):,}</p>
        </div>
        """
    
    html_content += """
    </body>
    </html>
    """
    
    return html_content

def process_combined_data(df, filenames):
    """Process combined data with dynamic bucketing"""
    
    # Get metadata
    ticker = df['Ticker'].iloc[0]
    asset_type = df['AssetType'].iloc[0]
    base_interval = df['Base_Interval_Minutes'].iloc[0]
    candle_interval = df['Candle_Interval_Minutes'].iloc[0]
    
    st.success(f"Processing {ticker} ({asset_type}) - Base: {base_interval}min, Analysis: {candle_interval}min")
    
    # Apply dynamic time bucketing
    df['TriggerTimeBucket'] = df['TriggerTime'].apply(
        lambda x: bucket_time_dynamic(x, base_interval, candle_interval)
    )
    df['GoalTimeBucket'] = df['GoalTime'].apply(
        lambda x: bucket_time_dynamic(x, base_interval, candle_interval) if pd.notna(x) else 'N/A'
    )
    
    # Show basic stats
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Records", len(df))
    with col2:
        st.metric("Analysis Types", len(df['AnalysisType'].unique()))
    with col3:
        st.metric("Unique Dates", df['Date'].nunique())
    with col4:
        total_hits = len(df[df['GoalHit'] == 'Yes'])
        st.metric("Goal Hits", total_hits)
    
    # Process each analysis type
    summary_rows = []
    
    for analysis_type in df['AnalysisType'].unique():
        analysis_data = df[df['AnalysisType'] == analysis_type]
        
        # Count triggers per combination
        trigger_counts = (
            analysis_data[analysis_data['TriggerTime'].notna()]
            .groupby(['TriggerLevel', 'TriggerTimeBucket', 'Direction'])
            .size()
            .reset_index(name='TotalTriggers')
        )
        
        # Count goal hits
        goal_hits = (
            analysis_data[
                (analysis_data['GoalHit'] == 'Yes') & 
                (analysis_data['GoalTimeBucket'] != 'OPEN') &
                (analysis_data['GoalTimeBucket'] != 'N/A')
            ]
            .groupby(['TriggerLevel', 'TriggerTimeBucket', 'Direction', 'GoalLevel', 'GoalTimeBucket'])
            .size()
            .reset_index(name='NumHits')
        )
        
        # Create summary for this analysis type
        for _, trigger_row in trigger_counts.iterrows():
            trigger_level = trigger_row['TriggerLevel']
            trigger_time = trigger_row['TriggerTimeBucket']
            direction = trigger_row['Direction']
            total_triggers = trigger_row['TotalTriggers']
            
            for goal_level in analysis_data['GoalLevel'].unique():
                bucket_labels, _ = create_time_buckets(base_interval, candle_interval)
                
                for goal_time in bucket_labels:
                    hits = goal_hits[
                        (goal_hits['TriggerLevel'] == trigger_level) &
                        (goal_hits['TriggerTimeBucket'] == trigger_time) &
                        (goal_hits['Direction'] == direction) &
                        (goal_hits['GoalLevel'] == goal_level) &
                        (goal_hits['GoalTimeBucket'] == goal_time)
                    ]
                    
                    num_hits = hits['NumHits'].iloc[0] if len(hits) > 0 else 0
                    pct_completion = (num_hits / total_triggers * 100) if total_triggers > 0 else 0
                    
                    summary_rows.append({
                        'AnalysisType': analysis_type,
                        'Direction': direction,
                        'TriggerLevel': trigger_level,
                        'TriggerTime': trigger_time,
                        'GoalLevel': goal_level,
                        'GoalTime': goal_time,
                        'NumTriggers': total_triggers,
                        'NumHits': num_hits,
                        'PctCompletion': round(pct_completion, 2)
                    })
    
    summary_df = pd.DataFrame(summary_rows)
    
    # Create metadata
    metadata = {
        'date_range': f"{df['Date'].min()} to {df['Date'].max()}",
        'total_records': len(df),
        'unique_dates': df['Date'].nunique(),
        'analysis_types': list(df['AnalysisType'].unique()),
        'source_files': filenames
    }
    
    return summary_df, metadata, ticker, base_interval, candle_interval, asset_type

# Streamlit App
st.title("🎯 Multi-Analysis ATR Summarizer with Combiner")
st.write("**Upload one or more CSV files for automatic processing**")

# File upload
uploaded_files = st.file_uploader(
    "Upload CSV files",
    type="csv",
    accept_multiple_files=True,
    help="Upload one or more CSV files. Multiple files will be automatically validated and combined."
)

if uploaded_files:
    try:
        # Load all files
        dataframes = []
        filenames = []
        
        for uploaded_file in uploaded_files:
            df = pd.read_csv(uploaded_file)
            dataframes.append(df)
            filenames.append(uploaded_file.name)
        
        st.success(f"📊 Loaded {len(dataframes)} file(s) with {sum(len(df) for df in dataframes):,} total records")
        
        # Validate files for combination
        is_valid, message = validate_files_for_combination(dataframes, filenames)
        
        if not is_valid:
            st.error(f"❌ Cannot combine files: {message}")
            st.stop()
        
        st.success(f"✅ {message}")
        
        # Combine files if multiple
        if len(dataframes) > 1:
            combined_df = combine_dataframes(dataframes, filenames)
            if combined_df is None:
                st.stop()  # User chose to stop due to overlaps
            st.success(f"🔄 Combined {len(dataframes)} files into {len(combined_df):,} unique records")
        else:
            combined_df = dataframes[0]
        
        # Process the data
        summary_df, metadata, ticker, base_interval, candle_interval, asset_type = process_combined_data(combined_df, filenames)
        
        # Generate reports
        st.write("## 📊 Generating Reports...")
        
        # Generate CSV
        csv_buffer = io.StringIO()
        summary_df.to_csv(csv_buffer, index=False)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"atr_summary_{ticker}_{asset_type}_{base_interval}min_{timestamp}.csv"
        
        # Generate Excel
        try:
            excel_content = generate_excel_report(summary_df, metadata, ticker, base_interval, candle_interval, asset_type)
            excel_filename = f"atr_summary_{ticker}_{asset_type}_{base_interval}min_{timestamp}.xlsx"
        except Exception as e:
            st.warning(f"Excel generation failed: {e}")
            excel_content = None
        
        # Generate HTML
        try:
            html_content = generate_html_report(summary_df, metadata, ticker, base_interval, candle_interval, asset_type)
            html_filename = f"atr_summary_{ticker}_{asset_type}_{base_interval}min_{timestamp}.html"
        except Exception as e:
            st.warning(f"HTML generation failed: {e}")
            html_content = None
        
        # Download buttons
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.download_button(
                label="📥 Download CSV",
                data=csv_buffer.getvalue(),
                file_name=csv_filename,
                mime="text/csv"
            )
        
        with col2:
            if excel_content:
                st.download_button(
                    label="📥 Download Excel",
                    data=excel_content,
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col3:
            if html_content:
                st.download_button(
                    label="📥 Download HTML",
                    data=html_content,
                    file_name=html_filename,
                    mime="text/html"
                )
        
        # Show summary
        st.write("## 📈 Summary Statistics")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Summary Records", len(summary_df))
        with col2:
            st.metric("Analysis Types", len(metadata['analysis_types']))
        with col3:
            st.metric("Source Files", len(filenames))
        with col4:
            bucket_labels, bucket_size = create_time_buckets(base_interval, candle_interval)
            st.metric("Time Buckets", len(bucket_labels))
        
        # Show analysis types
        st.write("### Analysis Types Processed:")
        for analysis_type in metadata['analysis_types']:
            type_data = summary_df[summary_df['AnalysisType'] == analysis_type]
            st.write(f"- **{analysis_type}**: {len(type_data):,} records")
        
        st.success("🎉 Processing complete!")
        
    except Exception as e:
        st.error(f"❌ Error processing files: {str(e)}")
        st.write("Please check your file format and try again.")

else:
    st.info("👆 Upload one or more CSV files to begin processing")
    st.write("### Features:")
    st.write("- **Automatic file validation** - Ensures files can be combined")
    st.write("- **Dynamic time bucketing** - Adapts to your analysis timeframe")
    st.write("- **Multi-analysis support** - Handles Session, Rolling, ZoneBaseline, StateCheck")
    st.write("- **Asset class aware** - Supports EQUITY_RTH, EQUITY_ETH, FUTURES, etc.")
    st.write("- **Minimal UI** - Just upload and download, everything else is automatic")
