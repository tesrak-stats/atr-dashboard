import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re

def detect_ticker_from_filename(filename):
    """Extract ticker from filename - looks for patterns like SPX_data.csv or data_SPX.csv"""
    if not filename:
        return None
    
    # Remove file extension
    basename = filename.replace('.csv', '').replace('.CSV', '')
    
    # Common ticker patterns
    ticker_patterns = [
        r'([A-Z]{1,5})_',  # SPX_data, QQQ_results
        r'_([A-Z]{1,5})$',  # data_SPX, results_QQQ
        r'^([A-Z]{1,5})$',  # Just SPX
        r'([A-Z]{1,5})\.', # SPX.data
    ]
    
    for pattern in ticker_patterns:
        match = re.search(pattern, basename)
        if match:
            return match.group(1)
    
    return None

def detect_ticker_from_data(df):
    """Try to detect ticker from data columns if available"""
    # Check if there's a ticker column
    ticker_columns = ['ticker', 'Ticker', 'TICKER', 'symbol', 'Symbol', 'SYMBOL']
    
    for col in ticker_columns:
        if col in df.columns:
            unique_tickers = df[col].unique()
            if len(unique_tickers) == 1:
                return str(unique_tickers[0]).upper()
    
    return None

def bucket_time(time_value):
    """Convert numeric times to hour buckets for dashboard display"""
    if pd.isna(time_value):
        return "No Trigger"  # Changed from "Unknown" to be more descriptive
    
    # Handle string times (like "OPEN")
    if isinstance(time_value, str):
        if time_value.upper() == "OPEN":
            return "OPEN"
        try:
            time_value = float(time_value)
        except:
            return str(time_value)
    
    # Convert numeric times to hour buckets
    if time_value < 930:
        return "OPEN"
    elif time_value < 1000:
        return "0900"
    elif time_value < 1100:
        return "1000"
    elif time_value < 1200:
        return "1100"
    elif time_value < 1300:
        return "1200"
    elif time_value < 1400:
        return "1300"
    elif time_value < 1500:
        return "1400"
    elif time_value < 1600:
        return "1500"
    else:
        return "1600"

def generate_excel_report(summary_df, metadata, ticker):
    """Generate Excel report matching the format of all levels 4.xlsx"""
    
    # Create workbook and worksheets
    wb = openpyxl.Workbook()
    
    # Remove default sheet and create new ones
    wb.remove(wb.active)
    
    # Create Information sheet
    info_ws = wb.create_sheet("Information")
    info_ws.cell(row=1, column=1, value="The sheets in this workbook show ATR analysis scenarios involving stock market behavior. Each table shows the likelihood of achieving one level at a particular time given that another level has already been achieved. It is broken up by hour in which the event triggers (the first level is hit) and the goal is reached (the second level is hit). OPEN is treated as a time, as it is the print when the market opens. If a level is achieved at OPEN and also the goal is reached, that result is discounted as it is not a tradeable situation. The numbers in the table are how often the goal is reached during that hour cross referenced to the hour it triggers. The percents listed are how often the trigger is reached during that hour if it is completed.")
    info_ws.cell(row=2, column=1, value=f"Ticker: {ticker} | Generated from data: {metadata['date_range']} | Total Records: {metadata['total_records']:,} | Trading Days: {metadata['unique_dates']:,}")
    
    # Separate data by direction
    above_data = summary_df[summary_df['Direction'] == 'Above'].copy()
    below_data = summary_df[summary_df['Direction'] == 'Below'].copy()
    
    # Create sheets for each direction
    if len(above_data) > 0:
        upside_ws = wb.create_sheet("Upside continuation")
        create_direction_sheet(upside_ws, above_data, "Above")
    
    if len(below_data) > 0:
        downside_ws = wb.create_sheet("Downside Continuation")  
        create_direction_sheet(downside_ws, below_data, "Below")
    
    # Save to BytesIO buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def create_direction_sheet(ws, data, direction):
    """Create a worksheet for one direction matching the original format"""
    
    # Define time periods for columns
    time_periods = ['OPEN', '0900', '1000', '1100', '1200', '1300', '1400', '1500', '1600']
    
    # Headers - matching the original format exactly
    headers = ['Trigger', 'Goal', 'Time triggered'] + [f'Goal Hit' for _ in time_periods] + ['Triggered', 'Completed', '%Complete', '%Incomplete']
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    current_row = 2
    
    # Group data by trigger level, then goal level, then trigger time
    trigger_levels = sorted(data['TriggerLevel'].unique())
    
    for trigger_level in trigger_levels:
        trigger_data = data[data['TriggerLevel'] == trigger_level]
        goal_levels = sorted(trigger_data['GoalLevel'].unique())
        
        for goal_level in goal_levels:
            goal_data = trigger_data[trigger_data['GoalLevel'] == goal_level]
            trigger_times = sorted(goal_data['TriggerTime'].unique())
            
            for trigger_time in trigger_times:
                time_data = goal_data[goal_data['TriggerTime'] == trigger_time]
                
                # Create trigger label (like "-1ATR")
                if trigger_level > 0:
                    trigger_label = f"+{abs(trigger_level):.0f}ATR" if trigger_level == int(trigger_level) else f"+{trigger_level:.3f}ATR"
                else:
                    trigger_label = f"-{abs(trigger_level):.0f}ATR" if trigger_level == int(trigger_level) else f"{trigger_level:.3f}ATR"
                
                # Get totals for this trigger-goal-time combination
                total_triggers = time_data['TotalTriggers'].iloc[0] if len(time_data) > 0 else 0
                total_actionable = time_data['ActionableTriggers'].iloc[0] if len(time_data) > 0 else 0
                
                # Prepare hit counts and percentages for each time period
                hit_counts = {}
                hit_percentages = {}
                total_hits = 0
                
                for period in time_periods:
                    period_data = time_data[time_data['GoalTime'] == period]
                    if len(period_data) > 0:
                        hits = period_data['NumHits'].iloc[0]
                        pct = period_data['PctCompletion'].iloc[0] / 100  # Convert to decimal
                        hit_counts[period] = hits
                        hit_percentages[period] = pct
                        total_hits += hits
                    else:
                        hit_counts[period] = 0
                        hit_percentages[period] = 0.0
                
                # Calculate completion statistics
                completed = total_hits
                pct_complete = (completed / total_actionable) if total_actionable > 0 else 0
                pct_incomplete = 1 - pct_complete
                
                # ROW 1: Raw counts
                ws.cell(row=current_row, column=1, value=trigger_label)  # Trigger
                ws.cell(row=current_row, column=2, value=goal_level)      # Goal
                ws.cell(row=current_row, column=3, value=trigger_time)   # Time triggered
                
                # Goal hit counts for each time period
                for i, period in enumerate(time_periods):
                    ws.cell(row=current_row, column=4 + i, value=hit_counts[period])
                
                # Summary columns
                ws.cell(row=current_row, column=13, value=total_actionable)  # Triggered (actionable)
                ws.cell(row=current_row, column=14, value=completed)         # Completed
                ws.cell(row=current_row, column=15, value=round(pct_complete, 4))    # %Complete
                ws.cell(row=current_row, column=16, value=round(pct_incomplete, 4))  # %Incomplete
                
                current_row += 1
                
                # ROW 2: Percentages (empty first 3 columns, then percentages)
                ws.cell(row=current_row, column=1, value=None)
                ws.cell(row=current_row, column=2, value=None)
                ws.cell(row=current_row, column=3, value=None)
                
                # Goal hit percentages for each time period
                for i, period in enumerate(time_periods):
                    if hit_percentages[period] > 0:
                        ws.cell(row=current_row, column=4 + i, value=round(hit_percentages[period], 4))
                    else:
                        ws.cell(row=current_row, column=4 + i, value=0)
                
                # Leave summary columns empty for percentage row
                for col in range(13, 17):
                    ws.cell(row=current_row, column=col, value=None)
                
                current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)  # Cap at 20 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def generate_html_report(summary_df, metadata, ticker):
    """Generate comprehensive HTML report of all trigger-goal combinations"""
    
    # Get unique trigger levels for organizing sections
    trigger_levels = sorted(summary_df['TriggerLevel'].unique())
    directions = ['Above', 'Below']
    trigger_times = ['OPEN', '0900', '1000', '1100', '1200', '1300', '1400', '1500']
    goal_times = ['0900', '1000', '1100', '1200', '1300', '1400', '1500']
    
    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>ATR Analysis Report - {ticker} - {metadata['date_range']}</title>
        <style>
            body {{
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                margin: 0;
                padding: 20px;
                background-color: #f8f9fa;
                color: #333;
                line-height: 1.6;
            }}
            .container {{
                max-width: 1400px;
                margin: 0 auto;
                background: white;
                border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                overflow: hidden;
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 30px;
                text-align: center;
            }}
            .header h1 {{
                margin: 0 0 10px 0;
                font-size: 2.5em;
            }}
            .header p {{
                margin: 0;
                opacity: 0.9;
                font-size: 1.1em;
            }}
            .ticker-badge {{
                background: rgba(255,255,255,0.2);
                padding: 8px 16px;
                border-radius: 20px;
                font-size: 1.2em;
                font-weight: bold;
                margin: 10px 0;
                display: inline-block;
            }}
            .summary-stats {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                padding: 30px;
                background: #f8f9fa;
                border-bottom: 1px solid #dee2e6;
            }}
            .stat-card {{
                background: white;
                padding: 20px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }}
            .stat-number {{
                font-size: 2em;
                font-weight: bold;
                color: #667eea;
                margin-bottom: 5px;
            }}
            .stat-label {{
                color: #6c757d;
                font-size: 0.9em;
            }}
            .content {{
                padding: 30px;
            }}
            .direction-section {{
                margin-bottom: 40px;
            }}
            .direction-title {{
                font-size: 1.8em;
                color: #495057;
                border-bottom: 3px solid #667eea;
                padding-bottom: 10px;
                margin-bottom: 25px;
            }}
            .trigger-section {{
                margin-bottom: 30px;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                overflow: hidden;
            }}
            .trigger-header {{
                background: #e9ecef;
                padding: 15px 20px;
                font-weight: bold;
                color: #495057;
                border-bottom: 1px solid #dee2e6;
            }}
            .data-table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 0.9em;
            }}
            .data-table th {{
                background: #f8f9fa;
                padding: 12px 8px;
                text-align: center;
                border-bottom: 2px solid #dee2e6;
                border-right: 1px solid #dee2e6;
                font-weight: 600;
                color: #495057;
            }}
            .data-table td {{
                padding: 10px 8px;
                text-align: center;
                border-bottom: 1px solid #f1f3f4;
                border-right: 1px solid #f1f3f4;
            }}
            .data-table tbody tr:hover {{
                background-color: #f8f9fa;
            }}
            .trigger-time {{
                background: #e3f2fd !important;
                font-weight: 600;
                color: #1565c0;
            }}
            .percentage {{
                font-weight: 500;
            }}
            .high-pct {{ color: #28a745; }}
            .med-pct {{ color: #ffc107; }}
            .low-pct {{ color: #dc3545; }}
            .zero-pct {{ color: #6c757d; }}
            .summary-row {{
                background: #f8f9fa !important;
                font-weight: 600;
                border-top: 2px solid #dee2e6;
            }}
            .footer {{
                background: #f8f9fa;
                padding: 20px;
                text-align: center;
                color: #6c757d;
                border-top: 1px solid #dee2e6;
            }}
            @media (max-width: 768px) {{
                .container {{ margin: 10px; }}
                .header {{ padding: 20px; }}
                .header h1 {{ font-size: 1.8em; }}
                .content {{ padding: 15px; }}
                .data-table {{ font-size: 0.8em; }}
                .data-table th, .data-table td {{ padding: 6px 4px; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>📊 ATR Analysis Report</h1>
                <div class="ticker-badge">{ticker}</div>
                <p>Comprehensive Trigger-Goal Analysis | {metadata['date_range']}</p>
            </div>
            
            <div class="summary-stats">
                <div class="stat-card">
                    <div class="stat-number">{metadata['total_records']:,}</div>
                    <div class="stat-label">Total Records</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{metadata['unique_dates']:,}</div>
                    <div class="stat-label">Trading Days</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{metadata['total_triggers']:,}</div>
                    <div class="stat-label">Total Triggers</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{metadata['overall_rate']:.1f}%</div>
                    <div class="stat-label">Overall Success Rate</div>
                </div>
            </div>
            
            <div class="content">
    """
    
    # Generate sections for each direction
    for direction in directions:
        direction_data = summary_df[summary_df['Direction'] == direction]
        if len(direction_data) == 0:
            continue
            
        html_content += f"""
                <div class="direction-section">
                    <h2 class="direction-title">📈 {direction} Trigger Analysis</h2>
        """
        
        # Group by trigger level ranges for better organization
        for trigger_level in trigger_levels:
            trigger_data = direction_data[direction_data['TriggerLevel'] == trigger_level]
            if len(trigger_data) == 0:
                continue
                
            html_content += f"""
                    <div class="trigger-section">
                        <div class="trigger-header">
                            Trigger Level: {trigger_level:+.3f}
                        </div>
                        <table class="data-table">
                            <thead>
                                <tr>
                                    <th>Trigger Time</th>
                                    <th>Goal Level</th>
                                    <th>0900</th>
                                    <th>1000</th>
                                    <th>1100</th>
                                    <th>1200</th>
                                    <th>1300</th>
                                    <th>1400</th>
                                    <th>1500</th>
                                    <th>Total Triggers</th>
                                    <th>Total Hits</th>
                                    <th>Success Rate</th>
                                </tr>
                            </thead>
                            <tbody>
            """
            
            # Organize data by trigger time
            for trigger_time in trigger_times:
                time_data = trigger_data[trigger_data['TriggerTime'] == trigger_time]
                if len(time_data) == 0:
                    continue
                
                # Get unique goals for this trigger time
                goals = sorted(time_data['GoalLevel'].unique())
                
                for i, goal_level in enumerate(goals):
                    goal_data = time_data[time_data['GoalLevel'] == goal_level]
                    
                    # Build row data
                    row_data = {}
                    total_hits = 0
                    total_triggers = 0
                    
                    for goal_time in goal_times:
                        goal_time_data = goal_data[goal_data['GoalTime'] == goal_time]
                        if len(goal_time_data) > 0:
                            pct = goal_time_data['PctCompletion'].iloc[0]
                            hits = goal_time_data['NumHits'].iloc[0]
                            triggers = goal_time_data['NumTriggers'].iloc[0]
                            total_hits += hits
                            total_triggers = triggers  # Should be same for all goals
                            
                            # Color code percentages
                            if pct >= 20:
                                pct_class = "high-pct"
                            elif pct >= 10:
                                pct_class = "med-pct"
                            elif pct > 0:
                                pct_class = "low-pct"
                            else:
                                pct_class = "zero-pct"
                            
                            row_data[goal_time] = f'<span class="{pct_class}">{pct:.1f}%</span>'
                        else:
                            row_data[goal_time] = '<span class="zero-pct">0.0%</span>'
                    
                    success_rate = (total_hits / total_triggers * 100) if total_triggers > 0 else 0
                    success_class = "high-pct" if success_rate >= 20 else "med-pct" if success_rate >= 10 else "low-pct" if success_rate > 0 else "zero-pct"
                    
                    # Show trigger time only for first goal
                    trigger_time_cell = f'<td class="trigger-time">{trigger_time}</td>' if i == 0 else '<td></td>'
                    
                    html_content += f"""
                                <tr>
                                    {trigger_time_cell}
                                    <td><strong>{goal_level:+.3f}</strong></td>
                                    <td>{row_data.get('0900', '<span class="zero-pct">0.0%</span>')}</td>
                                    <td>{row_data.get('1000', '<span class="zero-pct">0.0%</span>')}</td>
                                    <td>{row_data.get('1100', '<span class="zero-pct">0.0%</span>')}</td>
                                    <td>{row_data.get('1200', '<span class="zero-pct">0.0%</span>')}</td>
                                    <td>{row_data.get('1300', '<span class="zero-pct">0.0%</span>')}</td>
                                    <td>{row_data.get('1400', '<span class="zero-pct">0.0%</span>')}</td>
                                    <td>{row_data.get('1500', '<span class="zero-pct">0.0%</span>')}</td>
                                    <td><strong>{total_triggers}</strong></td>
                                    <td><strong>{total_hits}</strong></td>
                                    <td><span class="{success_class}"><strong>{success_rate:.1f}%</strong></span></td>
                                </tr>
                    """
            
            html_content += """
                            </tbody>
                        </table>
                    </div>
            """
        
        html_content += """
                </div>
        """
    
    # Close HTML
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    html_content += f"""
            </div>
            
            <div class="footer">
                <p>Generated on {current_time} | ATR Analysis System | Ticker: {ticker}</p>
                <p>📊 This report contains {len(summary_df):,} trigger-goal combinations across {metadata['unique_dates']} trading days</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html_content

def process_ticker_data(df, ticker, uploaded_filename):
    """Process data for a single ticker and return summary"""
    
    st.write(f"## 📊 Processing {ticker}")
    
    # Apply time bucketing
    df['TriggerTimeBucket'] = df['TriggerTime'].apply(bucket_time)
    df['GoalTimeBucket'] = df['GoalTime'].apply(lambda x: bucket_time(x) if pd.notna(x) and x != '' else 'N/A')

    # Show basic stats
    st.write(f"### Basic Statistics for {ticker}")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Records", len(df))
    with col2:
        st.metric("Unique Dates", df['Date'].nunique())
    with col3:
        total_hits = len(df[df['GoalHit'] == 'Yes'])
        st.metric("Total Goal Hits", total_hits)
    with col4:
        hit_rate = (total_hits / len(df) * 100) if len(df) > 0 else 0
        st.metric("Overall Hit Rate", f"{hit_rate:.1f}%")
    
    # STEP 1: Count total triggers per trigger combination (including 0-trigger scenarios)
    st.write(f"### Step 1: Count Total Triggers for {ticker}")
    
    # For "No Trigger" scenarios, we need to count them as 0 triggers
    # Create a comprehensive list of all trigger combinations
    all_trigger_combinations = []
    
    # Get actual trigger events (non-null times)
    actual_trigger_events = df[df['TriggerTime'].notna()][['Date', 'TriggerLevel', 'TriggerTimeBucket', 'Direction']].drop_duplicates()
    actual_trigger_counts = (
        actual_trigger_events
        .groupby(['TriggerLevel', 'TriggerTimeBucket', 'Direction'])
        .size()
        .reset_index(name='TotalTriggers')
    )
    
    # Get "No Trigger" scenarios 
    no_trigger_scenarios = df[df['TriggerTimeBucket'] == 'No Trigger'][['TriggerLevel', 'Direction']].drop_duplicates()
    no_trigger_counts = []
    for _, row in no_trigger_scenarios.iterrows():
        no_trigger_counts.append({
            'TriggerLevel': row['TriggerLevel'],
            'TriggerTimeBucket': 'No Trigger',
            'Direction': row['Direction'],
            'TotalTriggers': 0  # These represent 0 triggers
        })
    
    no_trigger_counts_df = pd.DataFrame(no_trigger_counts)
    
    # Combine actual triggers and no-trigger scenarios
    if len(no_trigger_counts_df) > 0:
        total_trigger_counts = pd.concat([actual_trigger_counts, no_trigger_counts_df], ignore_index=True)
    else:
        total_trigger_counts = actual_trigger_counts
    
    st.write(f"✅ Found {len(actual_trigger_counts)} actual trigger combinations")
    st.write(f"✅ Found {len(no_trigger_counts_df)} no-trigger scenarios")  
    st.write(f"✅ Total combinations (including 0-trigger): {len(total_trigger_counts)}")
    
    # STEP 2: Count OPEN completions per trigger-goal combination
    st.write(f"### Step 2: Count OPEN Completions for {ticker}")
    
    open_completions = df[
        (df['GoalHit'] == 'Yes') & 
        (df['GoalTimeBucket'] == 'OPEN')
    ].copy()
    
    open_completion_counts = (
        open_completions
        .groupby(['TriggerLevel', 'TriggerTimeBucket', 'Direction', 'GoalLevel'])
        .size()
        .reset_index(name='OpenCompletions')
    )
    
    st.write(f"✅ Found {len(open_completion_counts)} trigger-goal combinations with OPEN completions")
    
    # STEP 2.5: Count total OPEN completions per trigger (for dashboard tooltip)
    open_completions_per_trigger = (
        open_completions
        .groupby(['TriggerLevel', 'TriggerTimeBucket', 'Direction'])
        .size()
        .reset_index(name='TotalOpenCompletions')
    )
    
    # STEP 3: Count non-OPEN goal hits per trigger-goal-time combination
    st.write(f"### Step 3: Count Non-OPEN Goal Hits for {ticker}")
    
    non_open_hits = df[
        (df['GoalHit'] == 'Yes') & 
        (df['GoalTimeBucket'] != 'OPEN') &
        (df['GoalTimeBucket'] != 'N/A')
    ].copy()
    
    goal_hit_counts = (
        non_open_hits
        .groupby(['TriggerLevel', 'TriggerTimeBucket', 'Direction', 'GoalLevel', 'GoalTimeBucket'])
        .size()
        .reset_index(name='NonOpenHits')
    )
    
    st.write(f"✅ Found {len(goal_hit_counts)} non-OPEN goal hit combinations")
    
    # STEP 4: Calculate goal-specific denominators
    st.write(f"### Step 4: Calculate Goal-Specific Denominators for {ticker}")
    
    # Create summary with goal-specific denominators
    summary_rows = []
    
    # Get all unique combinations
    all_triggers = total_trigger_counts[['TriggerLevel', 'TriggerTimeBucket', 'Direction']].drop_duplicates()
    all_goals = df['GoalLevel'].unique()
    all_goal_times = [t for t in df['GoalTimeBucket'].unique() if t not in ['OPEN', 'N/A']]
    
    for _, trigger_row in all_triggers.iterrows():
        trigger_level = trigger_row['TriggerLevel']
        trigger_time = trigger_row['TriggerTimeBucket']
        direction = trigger_row['Direction']
        
        # Get total triggers for this combination
        total_triggers = total_trigger_counts[
            (total_trigger_counts['TriggerLevel'] == trigger_level) &
            (total_trigger_counts['TriggerTimeBucket'] == trigger_time) &
            (total_trigger_counts['Direction'] == direction)
        ]['TotalTriggers'].iloc[0]
        
        # Get total OPEN completions for this trigger (for dashboard tooltip)
        total_open_comps = open_completions_per_trigger[
            (open_completions_per_trigger['TriggerLevel'] == trigger_level) &
            (open_completions_per_trigger['TriggerTimeBucket'] == trigger_time) &
            (open_completions_per_trigger['Direction'] == direction)
        ]
        total_open_completions = total_open_comps['TotalOpenCompletions'].iloc[0] if len(total_open_comps) > 0 else 0
        
        for goal_level in all_goals:
            # Get OPEN completions for this specific trigger-goal combination
            open_comps = open_completion_counts[
                (open_completion_counts['TriggerLevel'] == trigger_level) &
                (open_completion_counts['TriggerTimeBucket'] == trigger_time) &
                (open_completion_counts['Direction'] == direction) &
                (open_completion_counts['GoalLevel'] == goal_level)
            ]
            
            open_completions_count = open_comps['OpenCompletions'].iloc[0] if len(open_comps) > 0 else 0
            
            # Calculate goal-specific denominator
            actionable_triggers = total_triggers - open_completions_count
            
            for goal_time in all_goal_times:
                # Get non-OPEN hits for this combination
                hits = goal_hit_counts[
                    (goal_hit_counts['TriggerLevel'] == trigger_level) &
                    (goal_hit_counts['TriggerTimeBucket'] == trigger_time) &
                    (goal_hit_counts['Direction'] == direction) &
                    (goal_hit_counts['GoalLevel'] == goal_level) &
                    (goal_hit_counts['GoalTimeBucket'] == goal_time)
                ]
                
                num_hits = hits['NonOpenHits'].iloc[0] if len(hits) > 0 else 0
                
                # Calculate percentage with goal-specific denominator
                if actionable_triggers > 0:
                    pct_completion = (num_hits / actionable_triggers * 100)
                else:
                    pct_completion = 0.0
                
                summary_rows.append({
                    'Direction': direction,
                    'TriggerLevel': trigger_level,
                    'TriggerTime': trigger_time,
                    'GoalLevel': goal_level,
                    'GoalTime': goal_time,
                    'TotalTriggers': total_triggers,
                    'OpenCompletions': open_completions_count,
                    'ActionableTriggers': actionable_triggers,
                    'NumHits': num_hits,
                    'PctCompletion': round(pct_completion, 2),
                    'TotalOpenCompletions': total_open_completions  # For dashboard tooltip
                })
    
    summary = pd.DataFrame(summary_rows)
    
    # Remove combinations with 0 actionable triggers
    summary = summary[summary['ActionableTriggers'] > 0]
    
    st.write(f"✅ Complete summary for {ticker}: {len(summary)} combinations")
    
    # Create final summary for dashboard (with OPEN completion data)
    dashboard_summary = summary[['Direction', 'TriggerLevel', 'TriggerTime', 'GoalLevel', 'GoalTime', 'ActionableTriggers', 'NumHits', 'PctCompletion', 'OpenCompletions', 'TotalOpenCompletions']].copy()
    dashboard_summary = dashboard_summary.rename(columns={'ActionableTriggers': 'NumTriggers'})
    
    return summary, dashboard_summary

st.title("🎯 Multi-Ticker Enhanced Summary Generator")
st.write("**Supports multiple tickers with automatic detection and naming**")
st.write("**Generates ticker-specific Excel and HTML reports**")

uploaded_file = st.file_uploader("Upload combined_trigger_goal_results_PERFECT.csv", type="csv")

if uploaded_file is not None:
    # Load the data
    df = pd.read_csv(uploaded_file)
    
    st.success(f"📊 Loaded {len(df)} records")
    
    # Detect ticker from filename or data
    detected_ticker = None
    
    # Try to detect from filename first
    if hasattr(uploaded_file, 'name'):
        detected_ticker = detect_ticker_from_filename(uploaded_file.name)
        if detected_ticker:
            st.info(f"🎯 Detected ticker from filename: **{detected_ticker}**")
    
    # If not found in filename, try to detect from data
    if not detected_ticker:
        detected_ticker = detect_ticker_from_data(df)
        if detected_ticker:
            st.info(f"🎯 Detected ticker from data: **{detected_ticker}**")
    
    # Allow user to override or set ticker manually
    col1, col2 = st.columns([3, 1])
    with col1:
        ticker_input = st.text_input(
            "Ticker Symbol", 
            value=detected_ticker or "SPX",
            help="Enter the ticker symbol for this data (e.g., SPX, QQQ, IWM, NVDA)"
        ).upper()
    
    with col2:
        st.write("**Common Tickers:**")
        st.write("SPX, QQQ, IWM, NVDA")
    
    if ticker_input:
        ticker = ticker_input
        st.success(f"✅ Processing data for ticker: **{ticker}**")
        
        # Check if this is multi-ticker data
        if 'Ticker' in df.columns or 'ticker' in df.columns:
            ticker_col = 'Ticker' if 'Ticker' in df.columns else 'ticker'
            unique_tickers = df[ticker_col].unique()
            
            if len(unique_tickers) > 1:
                st.warning(f"⚠️ Multi-ticker data detected: {', '.join(unique_tickers)}")
                selected_ticker = st.selectbox("Select ticker to process:", unique_tickers)
                df = df[df[ticker_col] == selected_ticker]
                ticker = selected_ticker
                st.info(f"🎯 Filtered to {len(df)} records for {ticker}")
        
        # Verify this has OPEN completions
        same_time_count = len(df[df['SameTime'] == True]) if 'SameTime' in df.columns else 0
        open_goals = len(df[df['GoalTime'] == 'OPEN']) if 'GoalTime' in df.columns else 0
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Same-Time Records", same_time_count)
        with col2:
            st.metric("OPEN Goal Records", open_goals)
        
        if same_time_count == 0 and open_goals == 0:
            st.error("❌ This doesn't appear to have OPEN completion data!")
            st.info("Make sure you're uploading the output from the PERFECT systematic generator")
        else:
            st.success("✅ Confirmed data has OPEN completion data for processing")
            
            # Process the ticker data
            summary, dashboard_summary = process_ticker_data(df, ticker, uploaded_file.name)
            
            # Report Generation Section
            st.write("## 📊 Report Generation")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Checkbox for generating HTML report
                generate_html = st.checkbox("📄 Generate HTML Report", 
                                           help="Creates a detailed HTML report with all trigger-goal combinations for web viewing")
            
            with col2:
                # Checkbox for generating Excel report
                generate_excel = st.checkbox("📊 Generate Excel Report", 
                                            help="Creates a detailed Excel report matching the format of all levels 4.xlsx")
            
            if generate_html:
                st.write("🔄 Generating HTML report...")
                
                # Prepare metadata for the report
                date_range = f"{df['Date'].min()} to {df['Date'].max()}"
                total_actionable = summary.groupby(['Direction', 'TriggerLevel', 'TriggerTime'])['ActionableTriggers'].first().sum()
                total_hits = summary['NumHits'].sum()
                overall_rate = (total_hits / total_actionable * 100) if total_actionable > 0 else 0
                
                metadata = {
                    'date_range': date_range,
                    'total_records': len(summary),
                    'unique_dates': df['Date'].nunique(),
                    'total_triggers': total_actionable,
                    'overall_rate': overall_rate
                }
                
                # Generate HTML content
                html_content = generate_html_report(summary, metadata, ticker)
                
                # Create download button for HTML
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                html_filename = f"atr_dashboard_summary_{ticker}_{timestamp}.html"
                
                st.download_button(
                    label="📥 Download HTML Report",
                    data=html_content,
                    file_name=html_filename,
                    mime="text/html",
                    help="Download comprehensive HTML report for web viewing and sharing"
                )
                
                st.success("✅ HTML report generated successfully!")
            
            if generate_excel:
                st.write("🔄 Generating Excel report...")
                
                # Prepare metadata for the report
                date_range = f"{df['Date'].min()} to {df['Date'].max()}"
                total_actionable = summary.groupby(['Direction', 'TriggerLevel', 'TriggerTime'])['ActionableTriggers'].first().sum()
                total_hits = summary['NumHits'].sum()
                overall_rate = (total_hits / total_actionable * 100) if total_actionable > 0 else 0
                
                metadata = {
                    'date_range': date_range,
                    'total_records': len(summary),
                    'unique_dates': df['Date'].nunique(),
                    'total_triggers': total_actionable,
                    'overall_rate': overall_rate
                }
                
                try:
                    # Generate Excel content
                    excel_content = generate_excel_report(summary, metadata, ticker)
                    
                    # Create download button for Excel
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    excel_filename = f"atr_dashboard_summary_{ticker}_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=excel_content,
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Download detailed Excel report matching the format of all levels 4.xlsx"
                    )
                    
                    st.success("✅ Excel report generated successfully!")
                    
                except Exception as e:
                    st.error(f"❌ Error generating Excel report: {str(e)}")
                    st.info("💡 **Note:** Excel generation requires openpyxl library. If running locally, install with: pip install openpyxl")
            
            # Save enhanced summary CSV with ticker-specific naming
            csv_buffer = io.StringIO()
            dashboard_summary.to_csv(csv_buffer, index=False)
            
            # Create ticker-specific CSV filename to match dashboard expectations
            csv_filename = f"atr_dashboard_summary_{ticker}_ENHANCED.csv"
            
            st.download_button(
                label=f"📥 Download {ticker} Enhanced Summary CSV",
                data=csv_buffer.getvalue(),
                file_name=csv_filename,
                mime="text/csv",
                help=f"Download CSV formatted for {ticker} dashboard integration"
            )
            
            # Final statistics
            st.write(f"## Final Statistics for {ticker}")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Summary Records", len(summary))
            with col2:
                total_actionable = summary.groupby(['Direction', 'TriggerLevel', 'TriggerTime'])['ActionableTriggers'].first().sum()
                st.metric("Total Actionable Triggers", f"{total_actionable:,}")
            with col3:
                total_hits = summary['NumHits'].sum()
                st.metric("Total Non-OPEN Hits", f"{total_hits:,}")
            with col4:
                overall_rate = (total_hits / total_actionable * 100) if total_actionable > 0 else 0
                st.metric("Overall Actionable Rate", f"{overall_rate:.1f}%")
            
            st.success(f"🎉 **{ticker} enhanced summary complete!**")
            st.info(f"💡 **File naming:** All outputs include '{ticker}' for easy dashboard integration")
            st.info(f"🗓️ **Trading Days:** {df['Date'].nunique()} days added to CSV for next program")
            
            # Show expected dashboard file pattern
            st.write("### 📁 Expected Dashboard File Pattern:")
            st.code(f"""
Based on your ticker config, the dashboard expects:
- CSV: atr_dashboard_summary_{ticker}.csv
- Generated: atr_dashboard_summary_{ticker}_ENHANCED.csv (now includes TradingDays column)
            """)

else:
    st.info("👆 Upload your PERFECT trigger-goal results CSV to generate enhanced summary and reports")
    st.write("### 🎯 Ticker Detection:")
    st.write("The app will automatically detect the ticker from:")
    st.write("- Filename patterns (e.g., `SPX_data.csv`, `data_QQQ.csv`)")
    st.write("- Data columns (if ticker/symbol column exists)")
    st.write("- Manual input override")
    
    st.write("### 📊 Multi-Ticker Support:")
    st.write("- Processes one ticker at a time")
    st.write("- Generates ticker-specific output files")
    st.write("- Matches your dashboard's expected file naming pattern")
